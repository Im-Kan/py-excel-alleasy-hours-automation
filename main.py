from openpyxl import load_workbook
from consts import obter_nome_mes
from glob import glob
import time_explorer
import alleasy
import sys
import calendar

args = sys.argv
print(args.__len__())
if args.__len__() < 3 or args.__len__() > 3:
    print("=> precisa colocar dps do py main.py:")
    print("=> o mês/ano (em numero) que voce quer preencher as hrs =>06/2023")
    print("=> seu nome")
    print("ex: $py main,py 6/2023 Gabriel")
    raise TypeError("Erro nos Argumentos")
try:
    mes = int(args[1].split("/")[0])
    ano = int(args[1].split("/")[1])
except IndexError:
    raise IndexError("Erro na data(Primeiro argumento)")

_, dias_no_mes = calendar.monthrange(ano, mes)
nome_mes = obter_nome_mes(mes)
nome = args[2]

devops_file = glob("./time_explorer_excel/*.xlsx")[0]
wk_devops = load_workbook(devops_file)
ws_devops = wk_devops.active

alleasy_file = glob("./alleasy_model_excel/*.xlsx")[0]
wk_alleasy = load_workbook(alleasy_file)
ws_alleasy = wk_alleasy.active

calendar_te = time_explorer.get_calendar_hashmap(ws_devops)

alleasy.set_data(ws_alleasy, mes, ano, dias_no_mes)
alleasy.set_horas(ws_alleasy, calendar_te, dias_no_mes)

wk_alleasy.save(nome_mes+'_Horas_'+nome+'.xlsx')
